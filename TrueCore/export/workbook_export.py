import time
import os
from openpyxl import load_workbook, Workbook


# -------------------------------------------------
# WORKBOOK LOCATION
# -------------------------------------------------

WORKBOOK_PATH = os.path.join(os.getcwd(), "TrueCore", "Outputs", "TrueValour Operations.xlsx")

INTEL_HEADERS = [
    "Packet Confidence",
    "Approval Probability",
    "Submission Readiness",
    "Workflow Queue",
    "Next Action",
    "Denial Risk",
    "Evidence Sufficiency",
    "Clinical Coherence",
    "Trust Score",
    "Policy Confidence",
]


# -------------------------------------------------
# CREATE WORKBOOK IF MISSING
# -------------------------------------------------

def create_workbook_if_missing(path):

    if os.path.exists(path):
        return

    try:

        os.makedirs(os.path.dirname(path), exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Patients"

        headers = [
            "ICN",
            "Patient Name",
            "DOB",
            "Authorization",
            "Ordering Doctor",
            "",
            "Treatment",
            "ICD Codes",
            "",
            "",
            "",
            "Packet Link"
        ]

        ws.append(headers)

        wb.save(path)

        print("Workbook created:", path)

    except Exception as e:
        print("Failed to create workbook:", e)


def ensure_workbook_headers(ws):

    expected_headers = {
        "A1": "ICN",
        "B1": "Patient Name",
        "C1": "DOB",
        "D1": "Authorization",
        "E1": "Ordering Doctor",
        "G1": "Treatment",
        "H1": "ICD Codes",
        "L1": "Packet Link",
    }

    for cell_ref, header in expected_headers.items():
        if ws[cell_ref].value in (None, ""):
            ws[cell_ref] = header

    start_column = 13  # M

    for offset, header in enumerate(INTEL_HEADERS):
        ws.cell(row=1, column=start_column + offset).value = header


# -------------------------------------------------
# EXPORT PATIENT
# -------------------------------------------------

def export_patient(fields, packet_path, intel_summary=None):

    # Ensure workbook exists
    create_workbook_if_missing(WORKBOOK_PATH)

    if not os.path.exists(WORKBOOK_PATH):
        return

    try:

        wb = load_workbook(WORKBOOK_PATH)
        ws = wb["Patients"]
        ensure_workbook_headers(ws)

        icn = fields.get("va_icn") or fields.get("icn")
        auth = fields.get("authorization_number")

        # Duplicate protection
        for r in range(2, ws.max_row + 1):

            if ws[f"A{r}"].value == icn or ws[f"D{r}"].value == auth:
                print("Duplicate patient detected. Skipping export.")
                return

        row = ws.max_row + 1

        codes = fields.get("icd_codes", [])

        ws[f"A{row}"] = icn
        ws[f"B{row}"] = fields.get("patient_name")
        ws[f"C{row}"] = fields.get("dob")
        ws[f"D{row}"] = auth
        ws[f"E{row}"] = fields.get("ordering_doctor")

        if any(code.upper().startswith("M54") for code in codes):
            ws[f"G{row}"] = "Fibrin"

        ws[f"H{row}"] = ", ".join(codes)

        cell = ws[f"L{row}"]
        cell.value = "Open Packet"
        cell.hyperlink = os.path.abspath(packet_path)
        cell.style = "Hyperlink"

        intel_summary = dict(intel_summary or {})
        intel_values = [
            intel_summary.get("packet_confidence"),
            intel_summary.get("approval_probability"),
            intel_summary.get("submission_readiness"),
            intel_summary.get("workflow_queue"),
            intel_summary.get("next_action"),
            intel_summary.get("denial_risk"),
            intel_summary.get("evidence_sufficiency"),
            intel_summary.get("clinical_coherence"),
            intel_summary.get("trust_score"),
            intel_summary.get("policy_confidence"),
        ]

        for offset, value in enumerate(intel_values, start=13):
            ws.cell(row=row, column=offset).value = value

        for attempt in range(3):

            try:
                wb.save(WORKBOOK_PATH)
                break

            except PermissionError:

                if attempt < 2:
                    time.sleep(2)
                else:
                    print("Excel workbook is open. Close it and retry export.")

    except Exception as e:

        print("Workbook export error:", e)
